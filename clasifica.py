"""Classify Mexican CFDI 4.0 bills into SAT-aligned folders + Excel report.

Usage:
    uv run python clasifica.py            # dry run, just write the Excel
    uv run python clasifica.py --apply    # actually move files
    uv run python clasifica.py --years 2025 2026 --apply

Categories:
    Auto      car/transport (combustibles, refacciones, transporte)
    AdqMerca  UsoCFDI=G01
    DedPers   UsoCFDI=D01..D10  (personal — should NOT appear here)
    Gastos    everything else (G03, P01, S01, …)

Layout produced:
    facturas/<YEAR>/<Mes>-<YEAR>/<YYMM>-<Cat>/<file>
    facturas/<YEAR>/sin_xml/<file>          # orphan PDFs grouped by year
    facturas/sin_xml/<file>                 # orphan PDFs with unknown year
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
FACTURAS = ROOT / "facturas"

NS = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "cfdi33": "http://www.sat.gob.mx/cfd/3",
    "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital",
}

MES_ABR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

DEDUCCION_PERSONAL = {f"D{i:02d}" for i in range(1, 11)}

# ClaveProdServ prefixes that mean automotive / transport spend.
AUTO_PREFIXES = (
    "1510",   # combustibles, lubricantes (Magna, Premium, Diesel, aceites)
    "2511",   # vehículos
    "2512",   # vehículos comerciales
    "2517",   # partes y accesorios para vehículos
    "7811",   # servicios de transporte de pasajeros
    "7818",   # servicios de mantto. y reparación de vehículos
    "8413",   # seguros de vehículos
)

# Emisor RFCs known to be 100% automotive (overrides product code).
AUTO_RFCS = {
    "PET040903DH1",   # Pemex (mayor)
}


# ─────────────────────────────────────────────────────────────────────────────
# Pairing
# ─────────────────────────────────────────────────────────────────────────────


def stem_key(path: Path) -> str:
    """Normalize a filename so XML and PDF pair up regardless of case / `(1)` suffix."""
    name = path.stem
    name = re.sub(r"\s*\(\d+\)\s*$", "", name)  # drop trailing " (1)" / "(2)"
    # Some PDFs are named `<xml-name>.xml.pdf` (the original XML name verbatim).
    if name.lower().endswith(".xml"):
        name = name[: -len(".xml")]
    return name.casefold()


@dataclass
class Pair:
    key: str
    xml: Path | None = None
    pdfs: list[Path] = field(default_factory=list)


def discover_pairs(root: Path) -> dict[str, Pair]:
    pairs: dict[str, Pair] = {}
    for p in sorted(root.rglob("*")):
        if not p.is_file():
            continue
        ext = p.suffix.lower()
        if ext not in (".xml", ".pdf"):
            continue
        key = stem_key(p)
        pair = pairs.setdefault(key, Pair(key=key))
        if ext == ".xml":
            if pair.xml is None:
                pair.xml = p
            else:
                # Two XMLs with same stem - rare; keep the larger / older path
                # but record by appending a suffix to the second key.
                alt_key = f"{key}__dup__{len(pairs)}"
                pairs[alt_key] = Pair(key=alt_key, xml=p)
        else:
            pair.pdfs.append(p)
    return pairs


# ─────────────────────────────────────────────────────────────────────────────
# XML parsing
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class Factura:
    pair_key: str
    xml_path: Path
    pdf_paths: list[Path]
    version: str = ""
    fecha: datetime | None = None
    folio: str = ""
    serie: str = ""
    forma_pago: str = ""
    metodo_pago: str = ""
    moneda: str = ""
    tipo_cambio: str = ""
    tipo_comprobante: str = ""
    subtotal: float = 0.0
    descuento: float = 0.0
    total: float = 0.0
    emisor_rfc: str = ""
    emisor_nombre: str = ""
    emisor_regimen: str = ""
    receptor_rfc: str = ""
    receptor_nombre: str = ""
    uso_cfdi: str = ""
    receptor_regimen: str = ""
    iva_trasladado: float = 0.0
    iva_retenido: float = 0.0
    isr_retenido: float = 0.0
    claves_prod_serv: list[str] = field(default_factory=list)
    descripciones: list[str] = field(default_factory=list)
    uuid: str = ""
    categoria: str = ""
    confianza: str = ""
    razon: str = ""
    parse_error: str = ""


def _to_float(s: str | None) -> float:
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_factura(xml_path: Path, pdf_paths: list[Path], pair_key: str) -> Factura:
    f = Factura(pair_key=pair_key, xml_path=xml_path, pdf_paths=pdf_paths)
    try:
        tree = ET.parse(xml_path)
    except ET.ParseError as exc:
        f.parse_error = f"XML malformed: {exc}"
        return f

    root = tree.getroot()
    tag = root.tag

    # Detect namespace (cfdi v4.0 vs v3.3 vs other complementos)
    if tag == f"{{{NS['cfdi']}}}Comprobante":
        ns = "cfdi"
    elif tag == f"{{{NS['cfdi33']}}}Comprobante":
        ns = "cfdi33"
    else:
        f.parse_error = f"Not a Comprobante (root tag: {tag})"
        return f

    cfdi_ns = NS[ns]
    a = root.attrib
    f.version = a.get("Version", a.get("version", ""))
    fecha_raw = a.get("Fecha", "")
    if fecha_raw:
        try:
            f.fecha = datetime.fromisoformat(fecha_raw)
        except ValueError:
            pass
    f.folio = a.get("Folio", "")
    f.serie = a.get("Serie", "")
    f.forma_pago = a.get("FormaPago", "")
    f.metodo_pago = a.get("MetodoPago", "")
    f.moneda = a.get("Moneda", "")
    f.tipo_cambio = a.get("TipoCambio", "")
    f.tipo_comprobante = a.get("TipoDeComprobante", "")
    f.subtotal = _to_float(a.get("SubTotal"))
    f.descuento = _to_float(a.get("Descuento"))
    f.total = _to_float(a.get("Total"))

    emisor = root.find(f"{{{cfdi_ns}}}Emisor")
    if emisor is not None:
        f.emisor_rfc = emisor.attrib.get("Rfc", "")
        f.emisor_nombre = emisor.attrib.get("Nombre", "")
        f.emisor_regimen = emisor.attrib.get("RegimenFiscal", "")

    receptor = root.find(f"{{{cfdi_ns}}}Receptor")
    if receptor is not None:
        f.receptor_rfc = receptor.attrib.get("Rfc", "")
        f.receptor_nombre = receptor.attrib.get("Nombre", "")
        f.uso_cfdi = receptor.attrib.get("UsoCFDI", "")
        f.receptor_regimen = receptor.attrib.get("RegimenFiscalReceptor", "")

    conceptos = root.find(f"{{{cfdi_ns}}}Conceptos")
    if conceptos is not None:
        for c in conceptos.findall(f"{{{cfdi_ns}}}Concepto"):
            f.claves_prod_serv.append(c.attrib.get("ClaveProdServ", ""))
            desc = c.attrib.get("Descripcion", "").strip()
            if desc:
                f.descripciones.append(desc)

    impuestos = root.find(f"{{{cfdi_ns}}}Impuestos")
    if impuestos is not None:
        f.iva_trasladado = _to_float(impuestos.attrib.get("TotalImpuestosTrasladados"))
        retenciones = impuestos.find(f"{{{cfdi_ns}}}Retenciones")
        if retenciones is not None:
            for r in retenciones.findall(f"{{{cfdi_ns}}}Retencion"):
                imp = r.attrib.get("Impuesto", "")
                amt = _to_float(r.attrib.get("Importe"))
                if imp == "002":
                    f.iva_retenido += amt
                elif imp == "001":
                    f.isr_retenido += amt

    complemento = root.find(f"{{{cfdi_ns}}}Complemento")
    if complemento is not None:
        tfd = complemento.find(f"{{{NS['tfd']}}}TimbreFiscalDigital")
        if tfd is not None:
            f.uuid = tfd.attrib.get("UUID", "")

    classify(f)
    return f


# ─────────────────────────────────────────────────────────────────────────────
# Classification
# ─────────────────────────────────────────────────────────────────────────────


def classify(f: Factura) -> None:
    if f.tipo_comprobante == "P":
        f.categoria, f.confianza, f.razon = "Pagos", "alta", "TipoDeComprobante=P (complemento de pagos)"
        return
    if f.tipo_comprobante == "N":
        f.categoria, f.confianza, f.razon = "Nomina", "alta", "TipoDeComprobante=N (nómina)"
        return

    if f.uso_cfdi in DEDUCCION_PERSONAL:
        f.categoria = "DedPers"
        f.confianza = "alta"
        f.razon = f"UsoCFDI={f.uso_cfdi} (deducción personal — revisar)"
        return

    auto_hits = []
    if f.emisor_rfc in AUTO_RFCS:
        auto_hits.append(f"emisor RFC {f.emisor_rfc}")
    for clave in f.claves_prod_serv:
        if clave.startswith(AUTO_PREFIXES):
            auto_hits.append(f"ClaveProdServ {clave}")
            break

    if auto_hits:
        f.categoria = "Auto"
        f.confianza = "alta"
        f.razon = "; ".join(auto_hits)
        return

    if f.uso_cfdi == "G01":
        f.categoria = "AdqMerca"
        f.confianza = "alta"
        f.razon = "UsoCFDI=G01"
        return

    f.categoria = "Gastos"
    f.confianza = "media" if f.uso_cfdi == "G03" else "baja"
    f.razon = f"UsoCFDI={f.uso_cfdi or '∅'} (general)"


# ─────────────────────────────────────────────────────────────────────────────
# Filing
# ─────────────────────────────────────────────────────────────────────────────


def target_dir_for(f: Factura) -> Path | None:
    if f.fecha is None:
        return None
    year = f.fecha.year
    mes_abr = MES_ABR[f.fecha.month - 1]
    yymm = f"{year % 100:02d}{f.fecha.month:02d}"
    return FACTURAS / str(year) / f"{mes_abr}-{year}" / f"{yymm}-{f.categoria}"


def safe_move(src: Path, dst_dir: Path) -> Path:
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / src.name
    if dst.exists() and dst.resolve() == src.resolve():
        return dst  # already there
    if dst.exists():
        # Same name, different file → suffix it
        stem, suffix = dst.stem, dst.suffix
        i = 1
        while True:
            candidate = dst_dir / f"{stem}__dup{i}{suffix}"
            if not candidate.exists():
                dst = candidate
                break
            i += 1
    shutil.move(str(src), str(dst))
    return dst


def file_pair(f: Factura, *, apply: bool, allowed_years: set[int]) -> tuple[str, list[str]]:
    actions: list[str] = []
    if f.fecha is None or f.fecha.year not in allowed_years:
        return ("year-skip", actions)
    target = target_dir_for(f)
    if target is None:
        return ("no-date", actions)
    if not apply:
        actions.append(f"would move {f.xml_path.name} -> {target.relative_to(ROOT)}")
        for pdf in f.pdf_paths:
            actions.append(f"would move {pdf.name} -> {target.relative_to(ROOT)}")
        return ("planned", actions)

    moved_xml = safe_move(f.xml_path, target)
    f.xml_path = moved_xml
    actions.append(f"moved {moved_xml.relative_to(ROOT)}")
    new_pdfs = []
    for pdf in f.pdf_paths:
        moved = safe_move(pdf, target)
        new_pdfs.append(moved)
        actions.append(f"moved {moved.relative_to(ROOT)}")
    f.pdf_paths = new_pdfs
    return ("moved", actions)


YYMM_CAT_RE = re.compile(r"^\d{4}-(Auto|AdqMerca|DedPers|Gastos|Pagos|Nomina)$")


def file_orphan_pdf(pdf: Path, *, apply: bool) -> tuple[str, str]:
    # Respect any pre-existing manual classification: if the file is already
    # inside a YYMM-Cat folder or a sin_xml folder, leave it there.
    parent_name = pdf.parent.name
    if YYMM_CAT_RE.match(parent_name) or parent_name == "sin_xml":
        return ("kept", f"kept {pdf.relative_to(ROOT)} (already in {parent_name})")
    # Try to guess year from the path itself or from mtime.
    year_match = re.search(r"\b(20[0-9]{2})\b", str(pdf))
    if year_match:
        year = year_match.group(1)
    else:
        year = str(datetime.fromtimestamp(pdf.stat().st_mtime).year)
    target = FACTURAS / year / "sin_xml"
    if not apply:
        return ("planned", f"would move {pdf.name} -> {target.relative_to(ROOT)}")
    moved = safe_move(pdf, target)
    return ("moved", f"moved {moved.relative_to(ROOT)}")


# ─────────────────────────────────────────────────────────────────────────────
# Excel output
# ─────────────────────────────────────────────────────────────────────────────


COLUMNS = [
    ("uuid", "UUID"),
    ("fecha", "Fecha"),
    ("anio", "Año"),
    ("mes", "Mes"),
    ("tipo_cmp", "Tipo Cmp."),
    ("emisor_rfc", "Emisor RFC"),
    ("emisor_nombre", "Emisor Nombre"),
    ("receptor_rfc", "Receptor RFC"),
    ("uso_cfdi", "Uso CFDI"),
    ("regimen_emisor", "Régimen Emisor"),
    ("forma_pago", "Forma Pago"),
    ("metodo_pago", "Método Pago"),
    ("moneda", "Moneda"),
    ("tipo_cambio", "Tipo Cambio"),
    ("subtotal", "Subtotal"),
    ("descuento", "Descuento"),
    ("iva_trasladado", "IVA Trasladado"),
    ("iva_retenido", "IVA Retenido"),
    ("isr_retenido", "ISR Retenido"),
    ("total", "Total"),
    ("claves_prod_serv", "Claves ProdServ"),
    ("descripcion", "Descripción"),
    ("folio", "Folio"),
    ("serie", "Serie"),
    ("categoria", "Categoría"),
    ("confianza", "Confianza"),
    ("razon", "Razón"),
    ("ruta_xml", "Ruta XML"),
    ("ruta_pdf", "Ruta PDF"),
]


def row_for(f: Factura) -> dict:
    return {
        "uuid": f.uuid,
        "fecha": f.fecha.isoformat() if f.fecha else "",
        "anio": f.fecha.year if f.fecha else "",
        "mes": f.fecha.month if f.fecha else "",
        "tipo_cmp": f.tipo_comprobante,
        "emisor_rfc": f.emisor_rfc,
        "emisor_nombre": f.emisor_nombre,
        "receptor_rfc": f.receptor_rfc,
        "uso_cfdi": f.uso_cfdi,
        "regimen_emisor": f.emisor_regimen,
        "forma_pago": f.forma_pago,
        "metodo_pago": f.metodo_pago,
        "moneda": f.moneda,
        "tipo_cambio": f.tipo_cambio,
        "subtotal": f.subtotal,
        "descuento": f.descuento,
        "iva_trasladado": f.iva_trasladado,
        "iva_retenido": f.iva_retenido,
        "isr_retenido": f.isr_retenido,
        "total": f.total,
        "claves_prod_serv": ", ".join(dict.fromkeys(f.claves_prod_serv)),
        "descripcion": " | ".join(f.descripciones)[:500],
        "folio": f.folio,
        "serie": f.serie,
        "categoria": f.categoria,
        "confianza": f.confianza,
        "razon": f.razon,
        "ruta_xml": str(f.xml_path.relative_to(ROOT)) if f.xml_path else "",
        "ruta_pdf": "; ".join(str(p.relative_to(ROOT)) for p in f.pdf_paths),
    }


def write_excel(facturas: list[Factura], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = [label for _, label in COLUMNS]
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = fill
    ws.freeze_panes = "A2"

    money_cols = {"subtotal", "descuento", "iva_trasladado", "iva_retenido", "isr_retenido", "total"}

    for f in facturas:
        row = row_for(f)
        ws.append([row[k] for k, _ in COLUMNS])

    for col_idx, (key, _) in enumerate(COLUMNS, start=1):
        if key in money_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

    # auto-width (cap at 50)
    for col_idx, (key, label) in enumerate(COLUMNS, start=1):
        max_len = len(label)
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, min(50, len(str(v))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Resumen pivot: categoria × año-mes → totals
    summary = wb.create_sheet("Resumen")
    pivot: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"subtotal": 0.0, "iva": 0.0, "total": 0.0, "n": 0})
    for f in facturas:
        if not f.fecha:
            continue
        key = (f.categoria or "Sin categoría", f"{f.fecha.year}-{f.fecha.month:02d}")
        pivot[key]["subtotal"] += f.subtotal
        pivot[key]["iva"] += f.iva_trasladado
        pivot[key]["total"] += f.total
        pivot[key]["n"] += 1

    summary.append(["Categoría", "Año-Mes", "# facturas", "Subtotal", "IVA", "Total"])
    for cell in summary[1]:
        cell.font = bold
        cell.fill = fill
    for (cat, ym), vals in sorted(pivot.items()):
        summary.append([cat, ym, vals["n"], vals["subtotal"], vals["iva"], vals["total"]])
    for col in (4, 5, 6):
        for row_idx in range(2, summary.max_row + 1):
            summary.cell(row=row_idx, column=col).number_format = "#,##0.00"
    for col_idx in range(1, 7):
        summary.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--apply", action="store_true", help="actually move files (default: dry-run)")
    ap.add_argument("--years", nargs="+", type=int, default=[2025, 2026], help="years to refile")
    ap.add_argument("--out", default="clasificacion.xlsx", help="output Excel filename")
    args = ap.parse_args()

    allowed_years = set(args.years)
    if not FACTURAS.exists():
        print(f"ERROR: {FACTURAS} not found", file=sys.stderr)
        return 1

    pairs = discover_pairs(FACTURAS)
    facturas: list[Factura] = []
    skipped: list[tuple[Path, str]] = []
    orphan_pdfs: list[Path] = []
    skipped_companion_pdfs: list[Path] = []

    for key, pair in pairs.items():
        if pair.xml is None:
            orphan_pdfs.extend(pair.pdfs)
            continue
        f = parse_factura(pair.xml, pair.pdfs, key)
        if f.parse_error:
            skipped.append((pair.xml, f.parse_error))
            # Treat the XML and any companion PDFs as orphans so they get filed
            # under sin_xml/ instead of being left scattered.
            orphan_pdfs.extend(pair.pdfs)
            skipped_companion_pdfs.append(pair.xml)  # XML routed via orphan path
            continue
        facturas.append(f)

    facturas.sort(key=lambda f: (f.fecha or datetime.min, f.emisor_nombre))

    moved_count = planned_count = year_skipped = 0
    for f in facturas:
        status, _ = file_pair(f, apply=args.apply, allowed_years=allowed_years)
        if status == "moved":
            moved_count += 1
        elif status == "planned":
            planned_count += 1
        elif status == "year-skip":
            year_skipped += 1

    # Orphan PDFs (and skipped XMLs) only get moved on --apply.
    orphan_moved = orphan_kept = orphan_planned = 0
    for path in orphan_pdfs + skipped_companion_pdfs:
        status, _ = file_orphan_pdf(path, apply=args.apply)
        if status == "moved":
            orphan_moved += 1
        elif status == "kept":
            orphan_kept += 1
        elif status == "planned":
            orphan_planned += 1

    out_path = ROOT / args.out
    write_excel(facturas, out_path)

    skipped_csv = ROOT / "skipped_xml.csv"
    with skipped_csv.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["xml_path", "razon"])
        for path, reason in skipped:
            w.writerow([str(path.relative_to(ROOT)), reason])

    pdf_sin_csv = ROOT / "pdf_sin_xml.csv"
    with pdf_sin_csv.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["pdf_path"])
        for pdf in orphan_pdfs:
            w.writerow([str(pdf.relative_to(ROOT))])

    cat_counts: dict[str, int] = defaultdict(int)
    for f in facturas:
        if f.fecha and f.fecha.year in allowed_years:
            cat_counts[f.categoria] += 1

    print()
    print(f"Discovered: {len(pairs)} pairs ({len(facturas)} parsed CFDIs, {len(skipped)} XML skipped, {len(orphan_pdfs)} orphan PDFs)")
    print(f"In scope (years {sorted(allowed_years)}): {sum(cat_counts.values())} bills")
    for cat in ("Auto", "AdqMerca", "DedPers", "Gastos", "Pagos", "Nomina"):
        if cat_counts.get(cat):
            print(f"   {cat:10s} {cat_counts[cat]:>4d}")
    print(f"Year-skipped (outside {sorted(allowed_years)}): {year_skipped}")
    if args.apply:
        print(f"Moved: {moved_count} bill pairs + {orphan_moved} orphan PDFs ({orphan_kept} kept in pre-existing folders)")
    else:
        print(f"Dry-run: {planned_count} bill pairs would be moved + {orphan_planned} orphan PDFs ({orphan_kept} would be kept in place)")
        print("(re-run with --apply to actually move files)")
    print(f"Excel: {out_path.relative_to(ROOT)}")
    print(f"PDFs sin XML: {pdf_sin_csv.relative_to(ROOT)}")
    print(f"XMLs omitidos: {skipped_csv.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
